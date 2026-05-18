"""
Timesheet Merger
================
从多个源文件（员工姓名.xlsx）提取考勤数据，
统一汇总输出为单一 output_YYYYMMDD_HHMM.xlsx。

输出固定 8 列：
  Date | Role/Type | Task Nature | Module | Hours | Ref ID | Staff ID | Name

使用方法：
  python timesheet_merger.py
  python timesheet_merger.py --src ./源文件目录 --out ./输出目录

Sheet 选取规则：
  1. 优先使用 wb.active（默认打开页）
  2. active 不可用时，fallback 到 sheetname == 'Timesheet'
  3. 其他 sheet 一律跳过

Staff ID / Name 扫描：
  - 在前 10 行内扫描，允许不在同一行
"""

import argparse
import warnings
from datetime import datetime, date, timedelta
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

warnings.filterwarnings("ignore", category=UserWarning)

# ─────────────────────────────────────────────
# 配置区（按需修改）
# ─────────────────────────────────────────────
DEFAULT_SRC_DIR = "./source_files"
DEFAULT_OUT_DIR = "./output"

# 输出固定列定义（顺序即为输出顺序）
OUTPUT_COLS = ["Date", "Role/Type", "Task Nature", "Module", "Hours", "Ref ID", "Staff ID", "Name"]


# ─────────────────────────────────────────────
# Step 1：扫描源文件目录
# ─────────────────────────────────────────────
def discover_files(src_dir: Path) -> list[Path]:
    """返回目录下所有 .xlsx 文件路径（排除临时文件）"""
    files = [
        f for f in src_dir.glob("*.xlsx")
        if not f.name.startswith("~$")
    ]
    if not files:
        raise FileNotFoundError(f"在 {src_dir} 未找到任何 .xlsx 文件")
    print(f"[发现] 共找到 {len(files)} 个源文件")
    return sorted(files)


# ─────────────────────────────────────────────
# Step 2：日期处理工具
# ─────────────────────────────────────────────
_TEXT_DATE_FMTS = [
    "%d-%b-%y",    # 6-Apr-26
    "%d-%b-%Y",    # 6-Apr-2026
    "%d/%m/%Y",    # 06/04/2026
    "%d/%m/%y",    # 06/04/26
    "%m/%d/%Y",    # 04/06/2026 美式
    "%Y-%m-%d",    # 2026-04-06 标准格式兜底
]

def _parse_date(val) -> str | None:
    """
    将各种来源的日期值统一转为 yyyy-mm-dd 字符串。
    返回 None 表示无效日期（如 'Total'），调用方应过滤该行。
    """
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=int(val))).strftime("%Y-%m-%d")
        except Exception:
            return None
    text = str(val).strip()
    if not text:
        return None
    if "total" in text.lower():
        return None
    for fmt in _TEXT_DATE_FMTS:
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


# ─────────────────────────────────────────────
# Step 3：从单个 sheet 提取数据
# ─────────────────────────────────────────────
def _find_header_row(ws) -> tuple[int, dict]:
    """
    在前 15 行里找含 'date' 和 'hours' 的表头行。
    返回 (行号, {列名小写: 列索引 1-based})
    """
    for row_idx in range(1, 16):
        row_vals = [str(c.value or "").strip().lower() for c in ws[row_idx]]
        if "date" in row_vals and "hours" in row_vals:
            col_map = {v: i + 1 for i, v in enumerate(row_vals) if v}
            return row_idx, col_map
    raise ValueError(f"在 sheet '{ws.title}' 的前 15 行未找到含 'Date' 和 'Hours' 的表头")


def _get_col(col_map: dict, *candidates) -> int | None:
    """从多个候选列名里找第一个存在的列索引"""
    for name in candidates:
        if name in col_map:
            return col_map[name]
    return None


def extract_sheet(ws, staff_id: str, emp_name: str) -> list[dict]:
    """
    从 sheet 提取有效行，只保留 8 个目标字段。
    有效行条件：Date 可解析为日期 且 Hours 为数字。
    """
    header_row, col_map = _find_header_row(ws)
    data_start = header_row + 1

    c_date  = _get_col(col_map, "date")
    c_role  = _get_col(col_map,
                        "role / type", "role/type", "role/type ", "role / type ")
    c_task  = _get_col(col_map,
                        "task nature", "task nature ")
    c_mod   = _get_col(col_map,
                        "module", "module ")
    c_hours = _get_col(col_map,
                        "hours", "hours ")
    c_refid = _get_col(col_map,
                        "ref id", "ref id (no need)", "ref/ticket #",
                        "ref id (mandatory for cr and ce, optional for production incident and ec ticket for now)",
                        "ref id \n(mandatory for cr and ce, optional for production incident and ec ticket for now)",
                        "ref id (no need) ")

    rows = []
    for row in ws.iter_rows(min_row=data_start, values_only=True):

        def g(col_idx):
            if col_idx is None:
                return None
            v = row[col_idx - 1] if col_idx <= len(row) else None
            if isinstance(v, str):
                v = v.strip()
                return None if v == "" else v
            return v

        raw_date  = g(c_date)
        raw_hours = g(c_hours)

        if raw_date is None or raw_hours is None:
            continue
        try:
            hours_val = float(raw_hours)
        except (TypeError, ValueError):
            continue
        parsed_date = _parse_date(raw_date)
        if parsed_date is None:
            continue

        rows.append({
            "Date":        parsed_date,
            "Role/Type":   g(c_role),
            "Task Nature": g(c_task),
            "Module":      g(c_mod),
            "Hours":       hours_val,
            "Ref ID":      g(c_refid),
            "Staff ID":    staff_id,
            "Name":        emp_name,
        })
    return rows


# ─────────────────────────────────────────────
# Step 4：解析单个源文件
# ─────────────────────────────────────────────
def parse_source(path: Path) -> list[dict]:
    """
    读取源文件，返回标准化记录列表。
    Sheet 选取：优先 active sheet，fallback 到 'Timesheet'。
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True, keep_links=False)

    # ── 选取目标 sheet ──
    active_ws = wb.active
    if active_ws is not None:
        ws = active_ws
        print(f"  [候选] 使用 active sheet：'{ws.title}'")
    else:
        ws = None
        for sname in wb.sheetnames:
            if sname.lower() == "timesheet":
                ws = wb[sname]
                print(f"  [候选] active 不可用，fallback 到 sheet：'{sname}'")
                break
        if ws is None:
            print(f"  [跳过] {path.name}：未找到可用 sheet")
            wb.close()
            return []

    # ── 隐藏 sheet 检测 ──
    try:
        if ws.sheet_state != "visible":
            print(f"  [跳过] sheet '{ws.title}'：隐藏 sheet，忽略")
            wb.close()
            return []
    except AttributeError:
        pass

    # ── 前 10 行扫描 Staff ID / Name（允许不同行） ──
    staff_id, emp_name = "", ""
    for scan_row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        row_vals = list(scan_row)
        for i, val in enumerate(row_vals):
            label = str(val or "").strip().lower()
            if not staff_id and label == "staff id" and i + 1 < len(row_vals):
                staff_id = str(row_vals[i + 1] or "").strip()
            if not emp_name and label == "name" and i + 1 < len(row_vals):
                emp_name = str(row_vals[i + 1] or "").strip()
        if staff_id and emp_name:
            break

    if not staff_id and not emp_name:
        print(f"  [跳过] {path.name} → sheet '{ws.title}'：未找到 Staff ID / Name")
        wb.close()
        return []

    # ── 提取数据 ──
    try:
        rows = extract_sheet(ws, staff_id, emp_name)
        print(f"  [读取] '{ws.title}' ({emp_name} / {staff_id})：{len(rows)} 行有效数据")
    except ValueError as e:
        print(f"  [警告] {path.name} → '{ws.title}'：{e}，已跳过")
        rows = []

    wb.close()
    return rows


# ─────────────────────────────────────────────
# Step 5：写入输出文件
# ─────────────────────────────────────────────
def write_output(records: list[dict], out_path: Path):
    """
    创建干净的 xlsx，写入固定 8 列数据。
    表头加粗 + 浅蓝底色，首行冻结，列宽固定。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"

    header_font  = Font(bold=True)
    header_fill  = PatternFill("solid", fgColor="BDD7EE")
    header_align = Alignment(horizontal="center", vertical="center")

    # 写表头
    for col_idx, col_name in enumerate(OUTPUT_COLS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    # 写数据
    for row_idx, rec in enumerate(records, 2):
        ws.cell(row_idx, 1).value = rec["Date"]
        ws.cell(row_idx, 2).value = rec["Role/Type"]
        ws.cell(row_idx, 3).value = rec["Task Nature"]
        ws.cell(row_idx, 4).value = rec["Module"]
        ws.cell(row_idx, 5).value = rec["Hours"]
        ws.cell(row_idx, 6).value = rec["Ref ID"]
        ws.cell(row_idx, 7).value = rec["Staff ID"]
        ws.cell(row_idx, 8).value = rec["Name"]

    # 列宽
    for col_idx, width in zip(range(1, 9), [14, 16, 18, 18, 8, 14, 14, 20]):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = width

    # 冻结首行
    ws.freeze_panes = "A2"

    wb.save(out_path)
    print(f"[输出] → {out_path}（{len(records)} 行）")


# ─────────────────────────────────────────────
# 主流程
# ─────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Timesheet Merger")
    parser.add_argument("--src", default=DEFAULT_SRC_DIR, help="源文件目录")
    parser.add_argument("--out", default=DEFAULT_OUT_DIR, help="输出目录")
    args = parser.parse_args()

    src_dir = Path(args.src)
    out_dir = Path(args.out)

    if not src_dir.exists():
        raise FileNotFoundError(f"源文件目录不存在：{src_dir}")
    out_dir.mkdir(parents=True, exist_ok=True)

    # Step 1：扫描
    files = discover_files(src_dir)

    # Step 2-4：逐文件解析
    all_records: list[dict] = []
    all_staff: dict[tuple, None] = {}

    for idx, path in enumerate(files, 1):
        print(f"\n[{idx}/{len(files)}] 处理：{path.name}")
        try:
            rows = parse_source(path)
            all_records.extend(rows)
            for r in rows:
                all_staff.setdefault((r["Staff ID"], r["Name"]), None)
        except Exception as e:
            print(f"  [错误] {path.name}：{e}，已跳过")

    # Step 5：写入
    print()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    if all_records:
        out_path = out_dir / f"output_{timestamp}.xlsx"
        write_output(all_records, out_path)
    else:
        print("[跳过] 无有效数据，未生成输出文件")

    # Step 6：统计汇总
    print("\n" + "─" * 50)
    print("📊 员工统计汇总")
    print("─" * 50)
    print(f"\n【output】共 {len(all_staff)} 人：")
    if all_staff:
        for i, (sid, name) in enumerate(all_staff.keys(), 1):
            print(f"  {i:>3}. {sid:<12} {name}")
    else:
        print("  （无数据）")
    print(f"\n{'─' * 50}")
    print(f"合计：{len(all_records)} 行 / {len(all_staff)} 人。")


if __name__ == "__main__":
    main()
